"""
Views for Academic Admission System
RESTful API with state machine enforcement and schema generation.
"""
from rest_framework import viewsets, status, views
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO

from .models import (
    Program, ProgramField, Admission, AdmissionState,
    AdmissionStateLog, AdmissionEvent, InternalNote,
    ContentPage, Achievement, GalleryItem, Enquiry, AnalyticEvent, Faculty
)
from .serializers import (
    ProgramSerializer, ProgramSummarySerializer, ProgramFieldSerializer,
    AdmissionListSerializer, AdmissionDetailSerializer,
    AdmissionCreateSerializer, AdmissionStepSerializer,
    AdmissionSubmitSerializer, StateTransitionSerializer,
    InternalNoteSerializer, ContentPageSerializer,
    AchievementSerializer, GalleryItemSerializer, EnquirySerializer,
    FacultySerializer
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ProgramViewSet(viewsets.ModelViewSet):
    """
    Program API - Dynamic schema generation for frontend.
    Returns program config including fields to render.
    """
    queryset = Program.objects.filter(is_active=True)
    serializer_class = ProgramSerializer
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ProgramSummarySerializer
        return ProgramSerializer
    
    @action(detail=True, methods=['get'])
    def schema(self, request, pk=None):
        """Return program schema for frontend rendering"""
        program = self.get_object()
        fields = program.fields.filter(is_visible=True).order_by('step', 'display_order')
        field_serializer = ProgramFieldSerializer(fields, many=True)
        
        return Response({
            'program': ProgramSummarySerializer(program).data,
            'steps': {
                step: list(fields.filter(step=step).values())
                for step in fields.values_list('step', flat=True).distinct()
            },
            'schema': field_serializer.data,
            'age_range': {
                'min': program.min_age,
                'max': program.max_age
            }
        })


class AdmissionViewSet(viewsets.ModelViewSet):
    """
    Admission API - State machine controlled.
    Frontend cannot skip, jump, or override steps.
    """
    queryset = Admission.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AdmissionListSerializer
        elif self.action == 'create':
            return AdmissionCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return AdmissionStepSerializer
        return AdmissionDetailSerializer
    
    def get_queryset(self):
        """Filter admissions based on user role"""
        queryset = Admission.objects.all()
        
        # Filter by state
        state = self.request.query_params.get('state')
        if state:
            queryset = queryset.filter(state=state)
        
        # Filter by program
        program = self.request.query_params.get('program')
        if program:
            queryset = queryset.filter(program__slug=program)
        
        # Filter by date range
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(application_number__icontains=search) |
                Q(phone__icontains=search)
            )
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Create new admission with step 1 data"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        admission = serializer.save()
        
        # Return full admission data
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def complete_step(self, request, pk=None):
        """Complete current step and advance"""
        admission = self.get_object()
        
        # Check if admission is in draft state
        if admission.state != AdmissionState.DRAFT:
            return Response(
                {'error': 'Cannot modify submitted admission'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(admission, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        # Return updated admission
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit the admission - final submission"""
        admission = self.get_object()
        
        # Validate all data is present
        serializer = AdmissionSubmitSerializer(admission)
        serializer.is_valid(raise_exception=True)
        
        admission = serializer.submit()
        
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data)
    
    @action(detail=True, methods=['get'])
    def status(self, request, pk=None):
        """Get current status including completed steps"""
        admission = self.get_object()
        
        return Response({
            'application_number': admission.application_number,
            'state': admission.state,
            'current_step': admission.current_step,
            'completed_steps': admission.completed_steps,
            'is_submitted': admission.state != AdmissionState.DRAFT
        })
    
    @action(detail=True, methods=['post'])
    def add_note(self, request, pk=None):
        """Add internal note to admission"""
        admission = self.get_object()
        serializer = InternalNoteSerializer(data={
            **request.data,
            'admission': admission.id
        })
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def transition(self, request, pk=None):
        """Admin: Transition state (approve/reject)"""
        admission = self.get_object()
        serializer = StateTransitionSerializer(
            admission,
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        
        admin_user = request.user.username if request.user.is_authenticated else 'admin'
        admission = serializer.transition(admin_user=admin_user)
        
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data)


class InternalNoteViewSet(viewsets.ModelViewSet):
    """Internal notes for admissions - staff only"""
    serializer_class = InternalNoteSerializer
    queryset = InternalNote.objects.all()
    
    def get_queryset(self):
        admission_id = self.kwargs.get('admission_pk')
        if admission_id:
            return InternalNote.objects.filter(admission_id=admission_id)
        return InternalNote.objects.all()


class ContentPageViewSet(viewsets.ReadOnlyModelViewSet):
    """Content pages API - read only"""
    queryset = ContentPage.objects.filter(is_published=True)
    serializer_class = ContentPageSerializer
    lookup_field = 'slug'
    
    def get_queryset(self):
        # Filter by visibility dates
        now = timezone.now()
        queryset = super().get_queryset()
        queryset = queryset.filter(
            Q(visible_from__isnull=True) | Q(visible_from__lte=now)
        ).filter(
            Q(visible_until__isnull=True) | Q(visible_until__gte=now)
        )
        return queryset


class AchievementViewSet(viewsets.ReadOnlyModelViewSet):
    """Achievements API - read only"""
    queryset = Achievement.objects.filter(is_visible=True)
    serializer_class = AchievementSerializer
    lookup_field = 'pk'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by date
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        return queryset


class GalleryViewSet(viewsets.ReadOnlyModelViewSet):
    """Gallery API - read only"""
    queryset = GalleryItem.objects.filter(is_visible=True)
    serializer_class = GalleryItemSerializer
    lookup_field = 'pk'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Return visible items ordered by display_order
        return queryset.order_by('display_order')
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get latest gallery items"""
        items = self.get_queryset()[:10]
        serializer = self.get_serializer(items, many=True)
        return Response(serializer.data)


class EnquiryViewSet(viewsets.ModelViewSet):
    """Enquiries API"""
    queryset = Enquiry.objects.all()
    serializer_class = EnquirySerializer
    
    def get_queryset(self):
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            self.queryset = self.queryset.filter(status=status_filter)
        
        # Filter by program
        program = self.request.query_params.get('program')
        if program:
            self.queryset = self.queryset.filter(
                Q(program_interest__slug=program) |
                Q(tagged_programs__contains=[program])
            )
        
        return self.queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """Create new enquiry"""
        # Auto-tag programs from form
        tagged = request.data.get('tagged_programs', [])
        program_interest = request.data.get('program_interest')
        
        if program_interest and str(program_interest) not in tagged:
            tagged.append(str(program_interest))
            request.data['tagged_programs'] = tagged
        
        return super().create(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update enquiry status"""
        enquiry = self.get_object()
        serializer = EnquiryStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        enquiry.status = serializer.validated_data['status']
        enquiry.assigned_to = serializer.validated_data.get('assigned_to', '')
        enquiry.follow_up_notes = serializer.validated_data.get('follow_up_notes', '')
        
        if enquiry.status == 'closed':
            enquiry.closed_at = timezone.now()
        
        enquiry.save()
        
        return Response(EnquirySerializer(enquiry).data)


class AnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """Analytics API - read only"""
    
    @action(detail=False, methods=['get'])
    def admissions(self, request):
        """Get admission analytics"""
        # Date range for analytics
        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)
        
        admissions = Admission.objects.filter(created_at__gte=since)
        
        # State distribution
        state_dist = dict(admissions.values('state').annotate(
            count=Count('id')
        ).values_list('state', 'count'))
        
        # Program distribution
        program_dist = dict(admissions.filter(
            program__isnull=False
        ).values('program__name').annotate(
            count=Count('id')
        ).values_list('program__name', 'count'))
        
        # Average time per step
        time_per_step = {}
        for step in ['1', '2', '3']:
            avg_time = admissions.filter(
                time_spent_per_step__has_key=step
            ).aggregate(
                avg=Avg(f'time_spent_per_step__{step}')
            )['avg']
            time_per_step[f'step_{step}'] = avg_time
        
        # Drop-off analysis
        drop_off = {}
        for step in [1, 2, 3]:
            completed = admissions.filter(completed_steps__contains=[step]).count()
            started = admissions.filter(current_step__gte=step).count()
            if started > 0:
                drop_off[f'step_{step}'] = {
                    'started': started,
                    'completed': completed,
                    'drop_off_rate': (started - completed) / started * 100
                }
        
        return Response({
            'period_days': days,
            'total_admissions': admissions.count(),
            'state_distribution': state_dist,
            'program_distribution': program_dist,
            'avg_time_per_step': time_per_step,
            'drop_off_analysis': drop_off
        })
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get dashboard summary"""
        today = timezone.now().date()
        today_start = timezone.make_aware(
            datetime.combine(today, datetime.min.time())
        )
        
        # Today's admissions
        today_count = Admission.objects.filter(
            created_at__gte=today_start
        ).count()
        
        # Pending review
        pending = Admission.objects.filter(
            state=AdmissionState.SUBMITTED
        ).count()
        
        # Recently submitted (last 7 days)
        week_ago = today - timedelta(days=7)
        recent = Admission.objects.filter(
            submitted_at__gte=week_ago
        ).count()
        
        # Status breakdown
        status_breakdown = dict(Admission.objects.values('state').annotate(
            count=Count('id')
        ).values_list('state', 'count'))
        
        # Program demand
        program_demand = dict(Admission.objects.filter(
            program__isnull=False
        ).values('program__name').annotate(
            count=Count('id')
        ).order_by('-count').values_list('program__name', 'count')[:5])
        
        return Response({
            'today': today_count,
            'pending_review': pending,
            'recent_submissions': recent,
            'status_breakdown': status_breakdown,
            'top_programs': program_demand
        })


class HealthCheckView(views.APIView):
    """Health check endpoint"""
    
    def get(self, request):
        return Response({
            'status': 'healthy',
            'timestamp': timezone.now(),
            'version': '1.0.0'
        })


class FacultyViewSet(viewsets.ReadOnlyModelViewSet):
    """Faculty API - read only"""
    queryset = Faculty.objects.filter(is_active=True)
    serializer_class = FacultySerializer
    lookup_field = 'pk'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.order_by('display_order')
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get latest faculty members"""
        members = self.get_queryset()[:10]
        serializer = self.get_serializer(members, many=True)
        return Response(serializer.data)


class AdmissionExportView(views.APIView):
    """Export admissions to Excel"""
    
    def get(self, request):
        """Export admissions as XLSX file"""
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'error': 'openpyxl is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Get filters from query params
        program_slug = request.query_params.get('program')
        status_filter = request.query_params.get('status')
        
        # Build queryset
        queryset = Admission.objects.select_related('program')
        
        if program_slug:
            queryset = queryset.filter(program__slug=program_slug)
        
        if status_filter:
            queryset = queryset.filter(state=status_filter)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admissions"
        
        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        # Write headers
        headers = [
            'Student Name',
            'Age',
            'Phone Number',
            'Full Address',
            'Parent / Guardian Name',
            'Program Name',
            'Admission Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_num, admission in enumerate(queryset, 2):
            # Build full address
            full_address = f"{admission.address_house_name}, {admission.address_place}, {admission.address_post_office} - {admission.address_pin_code}, {admission.address_district}, {admission.address_state}"
            
            row_data = [
                admission.name,
                admission.age_at_submission or admission.age,
                f"{admission.phone_country_code} {admission.phone}",
                full_address,
                admission.guardian_name,
                admission.program.name if admission.program else '',
                admission.get_state_display(),
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='left')
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        program_part = program_slug if program_slug else 'all'
        filename = f"admissions_{program_part}_{date_str}.xlsx"
        
        return Response(
            buffer.getvalue(),
            status=status.HTTP_200_OK,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )


